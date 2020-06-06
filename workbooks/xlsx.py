#!/usr/bin/python
import os, sys, logging
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
logging.basicConfig(filename='logfile.log', format='%(asctime)s %(message)s', level=logging.DEBUG)
try:
    import openpyxl as oxl
except ModuleNotFoundError:
    print("ExceptionERROR: Missing packages (required: os, sys, logging, openpyxl).")


class Read:
    def __init__(self, *args, **kwargs):
        # args[0] = full_file_name --  absolute path of a workbook
        # args[1] = BOOL of read_only
        # args[2] = BOOL of data_only
        # args[3] = worksheet -- INT sheet number in workbook

        try:
            # open workbook in readwrite mode if commanded and data_only modes
            self.open_wb(args[0], args[1], args[2])
        except KeyError:
            # open workbook in read_only and data_only modes
            self.open_wb(args[0], True, True)
        except:
            self.wb = None

        try:
            self.open_ws(int(args[3]))
        except:
            try:
                self.open_ws(0)
            except:
                self.ws = None

        self.max_col = int
        self.max_row = int

    def close_wb(self):
        try:
            self.wb.close()
        except:
            pass

    def col_increase_letter(self, letters, *step):
        # letters = STR - one or multiple letters of a column
        # step[0] = INT - number of letters to increase (optional, default = 1)
        # returns letter(s) / chr (alphabetically increased by step)
        try:
            step = step[0]
        except KeyError:
            step = 1
        chr_no = self.col_name_to_num(letters)
        chr_no += step
        return self.col_num_to_name(chr_no)

    def col_name_to_num(self, letters):
        # letters = STR of column name, e.g., 'AB'
        pow = 1
        col_int = 0
        for letter in letters[::-1]:
            col_int += (int(letter, 36) - 9) * pow
            pow *= 26
        return col_int + 64

    def col_num_to_name(self, col_int):
        col_int -= 64
        letters = ''
        while col_int:
            mod = (col_int - 1) % 26
            letters += chr(mod + 65)
            col_int = (col_int - 1) // 26
        return ''.join(reversed(letters))

    def lookup_value_in_column(self, col, val):
        # returns the first row number that contains a VALUE val in a COLUMN col
        # col = STR of column number, e.g., col = 'A'
        # val = STR/FLOAT/INT value to look up in column

        col_num = self.col_name_to_num(col) - 64
        for i in range(1, self.max_row):
            if str(self.ws.cell(row=i, column=col_num).value).lower() == str(val).lower():
                return i

    def lookup_value_in_row(self, row, val):
        # returns the first row number that contains a VALUE val in a ROW row
        # row = INT of row number, e.g., col = 'A' (min. 1!)
        # val = STR/FLOAT/INT value to look up in column

        for j in range(1, self.max_col):
            if str(self.ws.cell(row=row, column=j).value).lower() == str(val).lower():
                return self.col_num_to_name(j + 64)

    def open_wb(self, xlsx_name, *read_modes):
        # read_modes[0] = read only -- BOOL (if true: read only = TRUE)
        # read_modes[1] = data only -- BOOL (if true: data only = TRUE)
        try:
            self.wb = oxl.load_workbook(filename=xlsx_name, read_only=read_modes[0], data_only=read_modes[1])
        except KeyError:
            try:
                self.wb = oxl.load_workbook(filename=xlsx_name)
            except ValueError:
                self.wb = oxl.Workbook()
                logging.info("Created new workbook (" + str(xlsx_name).split("\\")[-1] + ").")

    def open_ws(self, worksheet):
        # worksheet = INT
        try:
            self.ws = self.wb.worksheets[worksheet]
        except KeyError:
            try:
                self.ws = self.wb[self.wb.sheetnames[0]]
                logging.info("WARNING: Defined worksheet not available (using ws no. 0 instead.")
            except KeyError:
                self.ws = []
                logging.info("ERROR: No worksheet available.")

    def read_cell(self, column, row):
        # column = CHR - cell column
        # row = INT - cell row
        # reads COLUMN / ROW cell
        try:
            cell_value = str(self.ws[str(column) + str(row)].value)
        except KeyError:
            cell_value = "None"
            logging.info("   * WARNING: Undefined cell " + str(column) + str(row))

        if not (cell_value == "None"):
            try:
                cell_value = float(cell_value)
            except ValueError:
                if cell_value.lower() == "inf":
                    cell_value = float(10 ** 10)
        return cell_value

    def read_one_column(self, column, start_row):
        # reads COLUMN beginning at START_ROW until it meets an empty cell
        # col = STR, e.g., col = "B"
        # start_row = INT
        # returns column as LIST
        data = []
        valid_content = True
        __row__ = start_row
        while valid_content:
            try:
                cell_value = str(self.ws[str(column) + str(__row__)].value)
            except KeyError:
                cell_value = "None"
                logging.info("   * WARNING: Undefined cell " + str(column) + str(__row__))

            if not(cell_value == "None"):
                try:
                    data.append(float(cell_value))
                except ValueError:
                    if cell_value.lower() == "inf":
                        cell_value = float(10**10)
                    data.append(cell_value)
                __row__ += 1
            else:
                valid_content = False
        return data

    def read_multiple_columns(self, start_row=0, columns="ABC"):
        return [self.ws["{}{}".format(column, row)].value for row in range(start_row, len(self.ws.rows) + 1) for column
                in columns]

    def read_matrix(self, start_col, start_row, *args, **kwargs):
        # reads matrix beginning at cell (start_col, start_row) until it meets an empty cell
        # start_row = INT (First Row = 1, not 0 as in a list!)
        # start_col = CHR, e.g., start_col = "B"
        # ENSURE (1) that the matrix to read is separated from other data in the workbook by empty cells
        # ENSURE (2) that the matrix in the workbook does not contain any empty cell
        
        matrix = []
        valid_content = True
        __row__ = start_row
        while valid_content:
            valid_content = self.test_cell_content(start_col, __row__)

            if valid_content:
                try:
                    matrix.append(self.read_row_str(__row__, start_col))
                except:
                    logging.info("   * WARNING: Undefined row data " + str(__row__))
                cc = 0
                for el in matrix[-1]:
                    try:
                        matrix[-1][int(cc)] = float(el)
                    except:
                        pass
                    cc += 1
            else:
                valid_content = False
            __row__ += 1
        return matrix

    def read_row_str(self, row, start_col, **kwargs):
        # reads ROW beginning at COL until it meets an empty cell
        # row = INT (First Row = 1, not 0 as in a list!)
        # start_col = CHR, e.g., start_col = "B"

        # parse optional arguments
        col_skip = kwargs.get("col_skip")
        end_col = kwargs.get("end_col")
        if_row = kwargs.get("if_row")

        if not col_skip:
            col_skip = 1
        if not end_col:
            # assumption: column XFA is the last column that can be handled within a spreadsheet
            end_col = "XFA"
        if not if_row:
            # row that tells if the content of the actual cell is relevant
            if_row = row

        str_data = []
        valid_content = True
        __col__ = start_col
        end_col_num = self.col_name_to_num(end_col)

        while valid_content:
            cell_string = str(self.ws[str(__col__) + str(row)].value)
            valid_content = self.test_cell_content(str(__col__), if_row)
            if valid_content:
                str_data.append(cell_string)

            # update column to read next relevant row entry
            __col_num__ = self.col_name_to_num(__col__)  # convert ascii-chr to int
            __col_num__ += col_skip                      # add col_skip

            if not (__col_num__ > end_col_num):
                __col__ = self.col_num_to_name(__col_num__)  # re-convert int to ascii-chr
            else:
                valid_content = False
        return str_data

    def set_max_col(self, col_number):
        self.max_col = int(col_number)

    def set_max_row(self, row_number):
        self.max_row = int(row_number)

    def test_cell_content(self, column, row):
        # column = CHR - cell column
        # row = INT - cell row
        # reads COLUMN / ROW cell
        if str(self.ws[column + str(row)].value) == "None":
            return False
        else:
            return True

    def __call__(self):
        print("Class Info: <type> = XLSX manipulation")


class Workbook(Read):
    def __init__(self, *args, **kwargs):
        # args[0] = STR full_path to workbook template
        # args[1] = INT worksheet number
        try:
            Read.__init__(self, args[0], False, False, args[1])
        except KeyError:
            Read.__init__(self)
            try:
                self.open_wb(args[0])
                try:
                    self.open_ws(args[1])
                except KeyError:
                    pass
            except TypeError:
                pass

        # color definitions -- more: https://www.computerhope.com/htmcolor.htm
        self.white = 'FFFFFFFF'

    def save_close_wb(self, full_file_path):
        try:
            logging.info("   * Saving as: " + full_file_path)
            self.wb.save(full_file_path)
            self.wb.close()
        except:
            logging.info("ERROR: Invalid file name or data.")

    def write_data2cell(self, column, row, value):
        # writes VALUE to cell COLUMN (e.g., 'A') and ROW (e.g., 1)
        try:
            self.ws[str(column) + str(row)].value = value
        except KeyError:
            logging.info("   * ERROR: Could not write value to CELL " + str(column) + str(row))

    def write_data2column(self, column, start_row, data_list):
        # writes COLUMN (STR, e.g., 'A') beginning at START_ROW (INT, e.g., 1)
        # data_list is a LIST object
        logging.info("   * Writing column data starting at " + str(column) + str(start_row) + " ...")
        __row__ = start_row
        for val in data_list:
            try:
                self.ws[str(column) + str(__row__)].value = val
            except KeyError:
                logging.info("   * WARNING: Could not write column entry: " + str(val))
            __row__ += 1

    def write_data2row(self, row_no, start_col, data_list, **kwargs):
        # writes ROW beginning at COL
        # row_no = INT of row number to write in
        # start_col = CHR, e.g., start_col = "B"
        # data_list = LIST of data to write

        # parse optional arguments
        col_skip = kwargs.get("col_skip")
        end_col = kwargs.get("end_col")

        if not col_skip:
            col_skip = 1
        if not end_col:
            # assumption: column XFA is the last column that can be handled within a spreadsheet
            end_col = "XFA"

        logging.info("   * Writing data row starting at " + str(start_col) + str(row_no) + " ...")
        __col__ = start_col
        end_col_num = self.col_name_to_num(end_col)

        for val in data_list:
            self.ws[str(__col__) + str(row_no)].value = val

            # update column to read next relevant row entry
            __col_num__ = self.col_name_to_num(__col__)  # convert ascii-chr to int
            __col_num__ += col_skip                      # add col_skip

            if not (__col_num__ > end_col_num):
                __col__ = self.col_num_to_name(__col_num__)  # re-convert int to ascii-chr
